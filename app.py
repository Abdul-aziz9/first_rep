import openpyxl as xl
from openpyxl.chart import BarChart, Reference 


wb = xl.load_workbook('first_rep/transactions.xlsx') #to access a file
sheet = wb['Sheet1'] #to access a sheet
for row in range(2, sheet.max_row + 1):#to loop though the sheet
    cell = sheet.cell(row , 3)#to aceess cell
    #print(cell.value) to print value on the cell 
    corrected_price = cell.value * 0.9 #created 9% of the initial value in each cell 
    corrected_price_cell = sheet.cell(row , 4) #row coming from the loop
    corrected_price_cell.value = corrected_price #adding the value to the cell

Value = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4 )

chart = BarChart()
chart.add_data(Value)
sheet.add_chart(chart, 'e5')
wb.save('transactions_corrected.xlsx')
